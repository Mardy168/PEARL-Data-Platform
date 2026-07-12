from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PREFERRED_COLUMNS = [
    "Date Collected", "Run Time Cambodia", "Published Date", "Crop", "Country", "Topic",
    "Title", "Summary", "Publisher", "Publisher Domain", "Source Type", "Source Name",
    "Language", "URL", "Canonical URL", "Google News URL", "Search Query", "Article ID",
    "Master Status", "Status",
]
COLUMN_MAP = {
    "date_collected": "Date Collected", "run_time_cambodia": "Run Time Cambodia",
    "crop": "Crop", "country": "Country", "topic": "Topic", "title": "Title",
    "Summary": "Summary", "publisher_name": "Publisher", "publisher_domain": "Publisher Domain",
    "source_type": "Source Type", "source_name": "Source Name", "language": "Language",
    "url": "URL", "canonical_url": "Canonical URL", "google_news_url": "Google News URL",
    "search_query": "Search Query", "article_id": "Article ID", "master_status": "Master Status",
    "status": "Status",
}


def prepare_excel_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy().rename(columns=COLUMN_MAP)
    for col in PREFERRED_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out[PREFERRED_COLUMNS]


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column in enumerate(ws.iter_cols(), 1):
        max_len = max((len(str(cell.value or "")) for cell in column), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 65)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_excel(df: pd.DataFrame, path: str, sheet_name: str = "News") -> str:
    # Excel worksheets support 1,048,576 rows including the header.
    if len(df) > 1_048_575:
        raise RuntimeError(f"Excel row limit exceeded: {len(df)} data rows")
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    prepare_excel_dataframe(df).to_excel(path, index=False, sheet_name=sheet_name)
    wb = load_workbook(path)
    _style_sheet(wb[sheet_name])
    wb.save(path)
    return path


def write_qa_workbook(metrics: dict, source_health: pd.DataFrame, path: str) -> str:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame([{"Metric": k, "Value": v} for k, v in metrics.items()]).to_excel(
            writer, index=False, sheet_name="Run Summary"
        )
        source_health.to_excel(writer, index=False, sheet_name="Source Health")
    wb = load_workbook(path)
    for ws in wb.worksheets:
        _style_sheet(ws)
    wb.save(path)
    return path
